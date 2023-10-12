import openpyxl

def get_row(path,sheetN):
    workbook = openpyxl.load_workbook(path)
    sheet=workbook[sheetN]
    return sheet.max_row
def get_column(path,sheetN):
    workbook = openpyxl.load_workbook(path)
    sheet=workbook[sheetN]
    return sheet.max_column
def read_excel(path,sheetN,rowN,columnN):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetN]
    return sheet.cell(row=rowN,column=columnN).value
def write_excel(path,sheetN,rowN,columnN,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetN]
    sheet.cell(row=rowN,column=columnN).value = data
    workbook.save(path)

